"""
Fleet-Wide Intervention Analysis Service.

Analyzes counterfactual explanations across all low-performing sites to identify
strategic interventions that could upgrade the entire portfolio. Instead of
individual site recommendations, this service identifies patterns like:

    "42 sites would benefit from Extended Hours Initiative"

Key Components:
- FleetAnalysisJob: Async job that generates and clusters counterfactuals
- InterventionCluster: Strategic intervention with ROI projections
- Excel export with 4-sheet executive report

Phase 5 of the explainability pipeline.
"""

import threading
import time
import uuid
import json
import numpy as np
import pandas as pd
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path
from datetime import datetime
import warnings

# Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    warnings.warn("openpyxl not installed. Excel export will be disabled.")

from site_scoring.explainability.counterfactuals import (
    CounterfactualGenerator,
    UpgradePathClusterer,
    ACTIONABLE_FEATURES,
    Counterfactual,
)
from site_scoring.explainability.tiers import TIER_LABELS, TIER_COLORS


@dataclass
class InterventionCluster:
    """A fleet-wide intervention identified from counterfactual clustering."""
    cluster_id: int
    name: str
    description: str
    n_sites: int
    pct_of_total: float
    primary_changes: List[Dict[str, Any]]  # {feature, direction, avg_magnitude}
    estimated_tier_shift: Dict[str, int]  # {from_tier: count}
    example_sites: List[str]
    estimated_success_rate: float = 0.75

    def to_dict(self) -> dict:
        return {
            'cluster_id': self.cluster_id,
            'name': self.name,
            'description': self.description,
            'n_sites': self.n_sites,
            'pct_of_total': round(self.pct_of_total * 100, 1),
            'primary_changes': self.primary_changes,
            'estimated_tier_shift': self.estimated_tier_shift,
            'example_sites': self.example_sites[:5],
            'estimated_success_rate': round(self.estimated_success_rate * 100, 1),
        }


@dataclass
class FleetAnalysisResult:
    """Complete result of fleet-wide analysis."""
    job_id: str
    status: str  # pending, running, completed, failed
    start_time: datetime
    end_time: Optional[datetime] = None

    # Input stats
    total_sites_analyzed: int = 0
    low_tier_sites: int = 0

    # Results
    interventions: List[InterventionCluster] = field(default_factory=list)
    tier_distribution_before: Dict[int, int] = field(default_factory=dict)
    tier_distribution_after: Dict[int, int] = field(default_factory=dict)

    # Progress tracking
    progress_pct: float = 0.0
    progress_message: str = ""
    error_message: Optional[str] = None

    # Site-level details
    site_counterfactuals: Dict[str, List[dict]] = field(default_factory=dict)
    site_cluster_assignments: Dict[str, int] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            'job_id': self.job_id,
            'status': self.status,
            'start_time': self.start_time.isoformat(),
            'end_time': self.end_time.isoformat() if self.end_time else None,
            'total_sites_analyzed': self.total_sites_analyzed,
            'low_tier_sites': self.low_tier_sites,
            'interventions': [i.to_dict() for i in self.interventions],
            'tier_distribution_before': self.tier_distribution_before,
            'tier_distribution_after': self.tier_distribution_after,
            'progress_pct': round(self.progress_pct, 1),
            'progress_message': self.progress_message,
            'error_message': self.error_message,
        }


# Global job storage (in production, use Redis or database)
_fleet_analysis_jobs: Dict[str, FleetAnalysisResult] = {}
_current_job_lock = threading.Lock()


class FleetAnalysisJob:
    """
    Async job for fleet-wide intervention analysis.

    This job:
    1. Identifies sites in Tier 3-4 (Review Required / Not Recommended)
    2. Generates counterfactuals for each site (what changes would upgrade them?)
    3. Clusters counterfactuals to find common intervention patterns
    4. Returns strategic recommendations with site counts

    Args:
        model: Sklearn-compatible model (or wrapped PyTorch)
        train_data: DataFrame with feature data for training
        feature_names: List of feature names
        continuous_features: List of continuous feature names
        calibrator: Fitted probability calibrator
        tier_classifier: Tier classifier instance
        output_dir: Directory to save results
    """

    def __init__(
        self,
        model: Any,
        train_data: pd.DataFrame,
        feature_names: List[str],
        continuous_features: List[str],
        calibrator: Any,
        tier_classifier: Any,
        output_dir: Path = None,
    ):
        self.model = model
        self.train_data = train_data
        self.feature_names = feature_names
        self.continuous_features = continuous_features
        self.calibrator = calibrator
        self.tier_classifier = tier_classifier
        self.output_dir = output_dir or Path("site_scoring/outputs/fleet_analysis")

        self._thread: Optional[threading.Thread] = None
        self._stop_flag = False

    def start(
        self,
        site_data: pd.DataFrame,
        site_ids: List[str],
        raw_probabilities: List[float],
        n_counterfactuals: int = 3,
        n_clusters: int = 5,
    ) -> str:
        """
        Start fleet analysis job in background thread.

        Args:
            site_data: DataFrame with features for all sites
            site_ids: List of site identifiers
            raw_probabilities: Model predictions for each site
            n_counterfactuals: Number of counterfactuals per site
            n_clusters: Number of intervention clusters to find

        Returns:
            job_id: Unique identifier for this job
        """
        job_id = str(uuid.uuid4())[:8]

        result = FleetAnalysisResult(
            job_id=job_id,
            status='pending',
            start_time=datetime.now(),
        )

        with _current_job_lock:
            _fleet_analysis_jobs[job_id] = result

        self._thread = threading.Thread(
            target=self._run_analysis,
            args=(job_id, site_data, site_ids, raw_probabilities, n_counterfactuals, n_clusters),
            daemon=True,
        )
        self._thread.start()

        return job_id

    def stop(self):
        """Signal the job to stop."""
        self._stop_flag = True

    def _run_analysis(
        self,
        job_id: str,
        site_data: pd.DataFrame,
        site_ids: List[str],
        raw_probabilities: List[float],
        n_counterfactuals: int,
        n_clusters: int,
    ):
        """Background thread that runs the full analysis."""
        result = _fleet_analysis_jobs[job_id]
        result.status = 'running'

        try:
            # Step 1: Identify low-tier sites (Tier 3-4)
            result.progress_message = "Identifying low-performing sites..."
            result.progress_pct = 5

            calibrated_probs = self.calibrator.calibrate(np.array(raw_probabilities))

            low_tier_indices = []
            tier_before = {1: 0, 2: 0, 3: 0, 4: 0}

            for i, prob in enumerate(calibrated_probs):
                tier = self.tier_classifier.classify(prob)
                tier_before[tier.tier] += 1
                if tier.tier >= 3:  # Review Required or Not Recommended
                    low_tier_indices.append(i)

            result.total_sites_analyzed = len(site_ids)
            result.low_tier_sites = len(low_tier_indices)
            result.tier_distribution_before = tier_before

            if len(low_tier_indices) == 0:
                result.progress_message = "No low-tier sites found!"
                result.status = 'completed'
                result.end_time = datetime.now()
                return

            result.progress_pct = 10

            # Step 2: Initialize counterfactual generator
            result.progress_message = "Initializing counterfactual generator..."

            try:
                cf_generator = CounterfactualGenerator(
                    model=self.model,
                    train_data=self.train_data,
                    feature_names=self.feature_names,
                    continuous_features=self.continuous_features,
                )
            except Exception as e:
                result.error_message = f"Failed to initialize counterfactual generator: {str(e)}"
                result.status = 'failed'
                result.end_time = datetime.now()
                return

            result.progress_pct = 15

            # Step 3: Generate counterfactuals for each low-tier site
            all_counterfactuals: Dict[int, List[Counterfactual]] = {}
            site_cf_map: Dict[str, List[dict]] = {}
            total_low_tier = len(low_tier_indices)

            for progress_idx, site_idx in enumerate(low_tier_indices):
                if self._stop_flag:
                    result.status = 'stopped'
                    result.end_time = datetime.now()
                    return

                site_id = site_ids[site_idx]
                site_features = site_data.iloc[[site_idx]].copy()

                result.progress_message = f"Generating counterfactuals for site {site_id} ({progress_idx + 1}/{total_low_tier})"
                result.progress_pct = 15 + (progress_idx / total_low_tier) * 60

                try:
                    cfs = cf_generator.generate(
                        site_features,
                        n_counterfactuals=n_counterfactuals,
                        desired_class=1,
                    )
                    if cfs:
                        all_counterfactuals[site_idx] = cfs
                        site_cf_map[site_id] = [cf.to_dict() for cf in cfs]
                except Exception as e:
                    # Log but continue with other sites
                    warnings.warn(f"Failed to generate counterfactuals for {site_id}: {e}")
                    continue

            result.site_counterfactuals = site_cf_map
            result.progress_pct = 75

            if len(all_counterfactuals) < n_clusters:
                n_clusters = max(1, len(all_counterfactuals) // 2)

            # Step 4: Cluster counterfactuals to find intervention patterns
            result.progress_message = "Clustering counterfactuals to identify interventions..."

            actionable_features = [f for f in ACTIONABLE_FEATURES if f in self.feature_names]

            clusterer = UpgradePathClusterer(n_clusters=n_clusters, min_cluster_size=2)
            clusterer.fit(all_counterfactuals, actionable_features)

            result.progress_pct = 85

            # Step 5: Extract intervention clusters
            result.progress_message = "Building intervention recommendations..."

            upgrade_paths = clusterer.get_upgrade_paths(len(low_tier_indices))

            # Convert to InterventionCluster objects with more detail
            interventions = []
            for path in upgrade_paths:
                # Parse primary changes into structured format
                primary_changes = []
                for change_str in path.primary_changes:
                    direction = 'increase' if 'Increase' in change_str else 'decrease'
                    feature = change_str.replace('Increase ', '').replace('Decrease ', '')
                    primary_changes.append({
                        'feature': feature,
                        'direction': direction,
                        'display': self._format_feature_change(feature, direction),
                    })

                # Get example site IDs
                example_site_ids = [site_ids[i] for i in path.example_sites if i < len(site_ids)]

                interventions.append(InterventionCluster(
                    cluster_id=path.cluster_id,
                    name=path.name,
                    description=path.description,
                    n_sites=path.n_sites_applicable,
                    pct_of_total=path.pct_of_portfolio,
                    primary_changes=primary_changes,
                    estimated_tier_shift={'from_tier_3': path.n_sites_applicable // 2, 'from_tier_4': path.n_sites_applicable // 2},
                    example_sites=example_site_ids,
                    estimated_success_rate=path.estimated_success_rate,
                ))

            result.interventions = interventions
            result.progress_pct = 95

            # Step 6: Estimate tier distribution after intervention
            result.progress_message = "Estimating post-intervention tier distribution..."
            result.tier_distribution_after = self._estimate_tier_shift(tier_before, interventions)

            # Save results
            self.output_dir.mkdir(parents=True, exist_ok=True)
            result_path = self.output_dir / f"fleet_analysis_{job_id}.json"
            with open(result_path, 'w') as f:
                json.dump(result.to_dict(), f, indent=2, default=str)

            result.progress_pct = 100
            result.progress_message = "Analysis complete!"
            result.status = 'completed'
            result.end_time = datetime.now()

        except Exception as e:
            result.error_message = str(e)
            result.status = 'failed'
            result.end_time = datetime.now()
            import traceback
            traceback.print_exc()

    def _format_feature_change(self, feature: str, direction: str) -> str:
        """Format feature change for executive display."""
        feature_display = {
            'c_emv_enabled_encoded': 'EMV Payment',
            'c_nfc_enabled_encoded': 'Contactless Payment',
            'c_open_24_hours_encoded': '24-Hour Operation',
            'c_vistar_programmatic_enabled_encoded': 'Programmatic Advertising',
            'c_walk_up_enabled_encoded': 'Walk-up Service',
            'experience_type': 'Experience Type',
            'hardware_type': 'Hardware Type',
        }

        display_name = feature_display.get(feature, feature.replace('_', ' ').title())

        if direction == 'increase':
            if 'encoded' in feature:
                return f"Enable {display_name}"
            return f"Increase {display_name}"
        else:
            return f"Reduce {display_name}"

    def _estimate_tier_shift(
        self,
        tier_before: Dict[int, int],
        interventions: List[InterventionCluster],
    ) -> Dict[int, int]:
        """
        Estimate tier distribution after implementing interventions.

        Uses a conservative estimate based on estimated success rates.
        """
        tier_after = tier_before.copy()

        # Total sites that could be upgraded
        total_upgradable = sum(i.n_sites for i in interventions)

        # Conservative estimate: 60% of estimated success rate
        for intervention in interventions:
            upgrades = int(intervention.n_sites * intervention.estimated_success_rate * 0.6)

            # Move from Tier 3/4 to Tier 1/2
            from_tier_3 = min(upgrades // 2, tier_after.get(3, 0))
            from_tier_4 = min(upgrades - from_tier_3, tier_after.get(4, 0))

            tier_after[3] = tier_after.get(3, 0) - from_tier_3
            tier_after[4] = tier_after.get(4, 0) - from_tier_4
            tier_after[2] = tier_after.get(2, 0) + from_tier_3  # Tier 3 -> Tier 2
            tier_after[1] = tier_after.get(1, 0) + from_tier_4  # Tier 4 -> Tier 1/2

        return tier_after


def get_fleet_analysis_status(job_id: str) -> Optional[Dict]:
    """Get status of a fleet analysis job."""
    result = _fleet_analysis_jobs.get(job_id)
    if result is None:
        return None
    return result.to_dict()


def export_fleet_analysis_to_excel(job_id: str, output_path: Path = None) -> Optional[Path]:
    """
    Export fleet analysis results to Excel with 4 sheets.

    Sheets:
    1. Executive Summary - High-level recommendations
    2. Intervention Details - Full intervention descriptions
    3. Site List - All analyzed sites with cluster assignments
    4. Tier Shift Analysis - Before/after tier projections

    Args:
        job_id: The job ID to export
        output_path: Optional output path (defaults to outputs/fleet_analysis/)

    Returns:
        Path to the created Excel file, or None if failed
    """
    if not OPENPYXL_AVAILABLE:
        warnings.warn("openpyxl not installed. Cannot export to Excel.")
        return None

    result = _fleet_analysis_jobs.get(job_id)
    if result is None or result.status != 'completed':
        return None

    if output_path is None:
        output_path = Path(f"site_scoring/outputs/fleet_analysis/fleet_analysis_{job_id}.xlsx")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ==================== Sheet 1: Executive Summary ====================
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"

    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    subheader_font = Font(bold=True, size=11)

    # Title
    ws_exec['A1'] = "Fleet-Wide Intervention Analysis Report"
    ws_exec['A1'].font = Font(bold=True, size=18)
    ws_exec.merge_cells('A1:E1')

    ws_exec['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_exec['A3'] = f"Total Sites Analyzed: {result.total_sites_analyzed}"
    ws_exec['A4'] = f"Low-Tier Sites Identified: {result.low_tier_sites}"

    # Summary table header
    row = 6
    headers = ['Intervention', 'Sites Affected', '% of Portfolio', 'Estimated Success', 'Primary Action']
    for col, header in enumerate(headers, 1):
        cell = ws_exec.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Intervention summary rows
    for intervention in result.interventions:
        row += 1
        ws_exec.cell(row=row, column=1, value=intervention.name)
        ws_exec.cell(row=row, column=2, value=intervention.n_sites)
        ws_exec.cell(row=row, column=3, value=f"{intervention.pct_of_total * 100:.1f}%")
        ws_exec.cell(row=row, column=4, value=f"{intervention.estimated_success_rate * 100:.0f}%")
        if intervention.primary_changes:
            ws_exec.cell(row=row, column=5, value=intervention.primary_changes[0].get('display', ''))

    # Auto-width columns
    for col in range(1, 6):
        ws_exec.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # ==================== Sheet 2: Intervention Details ====================
    ws_detail = wb.create_sheet("Intervention Details")

    row = 1
    for intervention in result.interventions:
        ws_detail.cell(row=row, column=1, value=intervention.name)
        ws_detail.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1

        ws_detail.cell(row=row, column=1, value="Description:")
        ws_detail.cell(row=row, column=2, value=intervention.description)
        row += 1

        ws_detail.cell(row=row, column=1, value="Sites Affected:")
        ws_detail.cell(row=row, column=2, value=intervention.n_sites)
        row += 1

        ws_detail.cell(row=row, column=1, value="Primary Changes:")
        changes_text = ", ".join(c.get('display', '') for c in intervention.primary_changes)
        ws_detail.cell(row=row, column=2, value=changes_text)
        row += 1

        ws_detail.cell(row=row, column=1, value="Example Sites:")
        ws_detail.cell(row=row, column=2, value=", ".join(intervention.example_sites))
        row += 2

    ws_detail.column_dimensions['A'].width = 20
    ws_detail.column_dimensions['B'].width = 60

    # ==================== Sheet 3: Site List ====================
    ws_sites = wb.create_sheet("Site List")

    # Headers
    site_headers = ['Site ID', 'Counterfactuals Generated', 'Top Change 1', 'Top Change 2']
    for col, header in enumerate(site_headers, 1):
        cell = ws_sites.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for site_id, cfs in result.site_counterfactuals.items():
        ws_sites.cell(row=row, column=1, value=site_id)
        ws_sites.cell(row=row, column=2, value=len(cfs))

        # Extract top changes from first counterfactual
        if cfs:
            changes = cfs[0].get('changes', {})
            change_list = list(changes.keys())[:2]
            if len(change_list) > 0:
                ws_sites.cell(row=row, column=3, value=change_list[0])
            if len(change_list) > 1:
                ws_sites.cell(row=row, column=4, value=change_list[1])
        row += 1

    for col in range(1, 5):
        ws_sites.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # ==================== Sheet 4: Tier Shift Analysis ====================
    ws_tier = wb.create_sheet("Tier Shift Analysis")

    ws_tier['A1'] = "Tier Distribution Analysis"
    ws_tier['A1'].font = Font(bold=True, size=14)

    # Before/After table
    tier_headers = ['Tier', 'Before', 'After', 'Change']
    for col, header in enumerate(tier_headers, 1):
        cell = ws_tier.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for tier_num in range(1, 5):
        row = 3 + tier_num
        before_count = result.tier_distribution_before.get(tier_num, 0)
        after_count = result.tier_distribution_after.get(tier_num, 0)
        change = after_count - before_count

        ws_tier.cell(row=row, column=1, value=f"Tier {tier_num} - {TIER_LABELS.get(tier_num, '')}")
        ws_tier.cell(row=row, column=2, value=before_count)
        ws_tier.cell(row=row, column=3, value=after_count)

        change_cell = ws_tier.cell(row=row, column=4, value=change)
        if change > 0:
            change_cell.font = Font(color="27AE60")  # Green for increase
        elif change < 0:
            change_cell.font = Font(color="E74C3C")  # Red for decrease

    for col in range(1, 5):
        ws_tier.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

    # Save workbook
    wb.save(output_path)
    return output_path


# Convenience function for API
def start_fleet_analysis(
    model: Any,
    train_data: pd.DataFrame,
    feature_names: List[str],
    continuous_features: List[str],
    calibrator: Any,
    tier_classifier: Any,
    site_data: pd.DataFrame,
    site_ids: List[str],
    raw_probabilities: List[float],
    n_counterfactuals: int = 3,
    n_clusters: int = 5,
    output_dir: Path = None,
) -> str:
    """
    Convenience function to start fleet analysis.

    Returns:
        job_id for tracking progress
    """
    job = FleetAnalysisJob(
        model=model,
        train_data=train_data,
        feature_names=feature_names,
        continuous_features=continuous_features,
        calibrator=calibrator,
        tier_classifier=tier_classifier,
        output_dir=output_dir,
    )

    return job.start(
        site_data=site_data,
        site_ids=site_ids,
        raw_probabilities=raw_probabilities,
        n_counterfactuals=n_counterfactuals,
        n_clusters=n_clusters,
    )
